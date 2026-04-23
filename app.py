"""
AI Sector Paper Trading Simulator
Flask backend with real Yahoo Finance data via yfinance
"""

from flask import Flask, render_template, jsonify, request, send_file
from flask_cors import CORS
import yfinance as yf
import pandas as pd
import numpy as np
import json
import os
from datetime import datetime, timedelta
import io
import xml.etree.ElementTree as ET
import urllib.request
import urllib.error
import ssl
import sqlite3
import certifi
from dotenv import load_dotenv
import config

load_dotenv()

# config.py takes precedence over .env for OPENAI_API_KEY
if config.OPENAI_API_KEY:
    os.environ["OPENAI_API_KEY"] = config.OPENAI_API_KEY

app = Flask(__name__)
CORS(app)

# ── Sentiment Cache ─────────────────────────────────────────────────────────
# Maps ticker -> {score: float 0-1, reasoning: str, updated_at: str}
sentiment_cache: dict = {}

# ── Database ─────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(config.DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Initialize SQLite DB and seed initial portfolio if empty."""
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS portfolio (
            id         INTEGER PRIMARY KEY,
            cash       REAL NOT NULL,
            initial    REAL NOT NULL,
            updated    TEXT NOT NULL,
            started_at TEXT NOT NULL DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS positions (
            ticker      TEXT PRIMARY KEY,
            name        TEXT,
            sector      TEXT,
            shares      INTEGER,
            entry_price REAL,
            high_water  REAL,
            date_opened TEXT
        );

        CREATE TABLE IF NOT EXISTS trades (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            date       TEXT,
            action     TEXT,
            ticker     TEXT,
            name       TEXT,
            sector     TEXT,
            price      REAL,
            shares     INTEGER,
            pnl        REAL,
            signal     TEXT,
            strategy   TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS daily_runs (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            date            TEXT UNIQUE,
            portfolio_value REAL,
            cash            REAL,
            initial_capital REAL,
            return_pct      REAL,
            trades_count    INTEGER,
            notes           TEXT,
            created_at      TEXT DEFAULT CURRENT_TIMESTAMP
        );
    """)
    # Migrate existing DBs that may not have started_at column yet
    try:
        conn.execute("ALTER TABLE portfolio ADD COLUMN started_at TEXT NOT NULL DEFAULT ''")
        conn.commit()
    except Exception:
        pass  # column already exists

    row = conn.execute("SELECT COUNT(*) FROM portfolio").fetchone()[0]
    if row == 0:
        today = datetime.now().isoformat()[:10]
        conn.execute(
            "INSERT INTO portfolio (cash, initial, updated, started_at) VALUES (?,?,?,?)",
            (config.INITIAL_CAPITAL, config.INITIAL_CAPITAL, datetime.now().isoformat(), today)
        )
    else:
        # Patch empty started_at from earliest trade date (migration for existing DBs)
        empty = conn.execute("SELECT started_at FROM portfolio WHERE id=1").fetchone()[0]
        if not empty:
            earliest = conn.execute("SELECT MIN(date) FROM trades").fetchone()[0]
            if earliest:
                conn.execute("UPDATE portfolio SET started_at=? WHERE id=1", (earliest,))
    conn.commit()
    conn.close()


def load_portfolio() -> dict:
    conn = get_db()
    pr = conn.execute("SELECT cash, initial, started_at FROM portfolio WHERE id=1").fetchone()
    pos_rows = conn.execute("SELECT * FROM positions").fetchall()
    conn.close()
    positions = {}
    for r in pos_rows:
        positions[r["ticker"]] = {
            "ticker":     r["ticker"],
            "name":       r["name"] or r["ticker"],
            "sector":     r["sector"] or "AI",
            "shares":     r["shares"],
            "entry":      r["entry_price"],
            "current":    r["entry_price"],
            "high_water": r["high_water"],
            "day_opened": r["date_opened"] or "",
        }
    return {"cash": pr["cash"], "initial": pr["initial"],
            "started_at": pr["started_at"] or "", "positions": positions}


def save_portfolio(portfolio: dict):
    conn = get_db()
    conn.execute(
        "UPDATE portfolio SET cash=?, updated=? WHERE id=1",
        (portfolio["cash"], datetime.now().isoformat())
    )
    conn.execute("DELETE FROM positions")
    for ticker, p in portfolio["positions"].items():
        conn.execute(
            """INSERT INTO positions
               (ticker, name, sector, shares, entry_price, high_water, date_opened)
               VALUES (?,?,?,?,?,?,?)""",
            (p["ticker"], p.get("name", ticker), p.get("sector", "AI"),
             p["shares"], p["entry"], p.get("high_water", p["entry"]),
             p.get("day_opened", datetime.now().isoformat()[:10]))
        )
    conn.commit()
    conn.close()


def save_trades(trades: list):
    if not trades:
        return
    conn = get_db()
    conn.executemany(
        """INSERT INTO trades
           (date, action, ticker, name, sector, price, shares, pnl, signal, strategy)
           VALUES (:date,:action,:ticker,:name,:sector,:price,:shares,:pnl,:signal,:strategy)""",
        trades
    )
    conn.commit()
    conn.close()


def save_daily_run(date: str, portfolio_value: float, cash: float, initial: float,
                   trades_count: int, notes: str):
    ret = (portfolio_value - initial) / initial * 100 if initial > 0 else 0
    conn = get_db()
    conn.execute(
        """INSERT INTO daily_runs
           (date, portfolio_value, cash, initial_capital, return_pct, trades_count, notes)
           VALUES (?,?,?,?,?,?,?)
           ON CONFLICT(date) DO UPDATE SET
             portfolio_value=excluded.portfolio_value,
             cash=excluded.cash,
             return_pct=excluded.return_pct,
             trades_count=trades_count + excluded.trades_count,
             notes=CASE WHEN notes='' THEN excluded.notes
                        ELSE notes || '; ' || excluded.notes END""",
        (date, portfolio_value, cash, initial, ret, trades_count, notes)
    )
    conn.commit()
    conn.close()


# ── Stock Universe ──────────────────────────────────────────────────────────

SECTORS = {
    "semis": {
        "name": "Semiconductors",
        "color": "#185FA5",
        "stocks": ["NVDA", "AMD", "AVGO", "QCOM", "MRVL", "TSM", "INTC", "AMAT"]
    },
    "cloud_ai": {
        "name": "Cloud & AI Infrastructure",
        "color": "#0F6E56",
        "stocks": ["MSFT", "GOOGL", "AMZN", "META", "ORCL", "IBM", "SNOW", "NET"]
    },
    "ai_software": {
        "name": "AI Software",
        "color": "#534AB7",
        "stocks": ["CRM", "PLTR", "AI", "PATH", "BBAI", "SOUN", "AMBA", "CIEN"]
    },
    "robotics": {
        "name": "Robotics & Automation",
        "color": "#993C1D",
        "stocks": ["ISRG", "ROK", "TER", "NDSN", "ONTO", "ACMR", "RRX", "AME"]
    },
    "ev_auto": {
        "name": "EV & Autonomous",
        "color": "#BA7517",
        "stocks": ["TSLA", "RIVN", "LCID", "MBLY", "BLNK", "CHPT", "LEA", "APTV"]
    },
    "biotech_ai": {
        "name": "AI Biotech",
        "color": "#993556",
        "stocks": ["ILMN", "RXRX", "SDGR", "BEAM", "CRSP", "EDIT", "NTLA", "PACB"]
    }
}

# ── Extended sector universe for the Stock Picker ────────────────────────────
# ALL_SECTORS = original 6 AI sectors + 9 additional popular industries
ALL_SECTORS = {
    **SECTORS,
    "fintech": {
        "name": "Fintech & Payments",
        "color": "#1A7A4A",
        "stocks": ["V", "MA", "PYPL", "SQ", "AFRM", "SOFI", "NU", "COIN"]
    },
    "cybersec": {
        "name": "Cybersecurity",
        "color": "#2C5282",
        "stocks": ["CRWD", "PANW", "ZS", "FTNT", "S", "OKTA", "CYBR", "QLYS"]
    },
    "healthcare": {
        "name": "Healthcare & MedTech",
        "color": "#6B3FA0",
        "stocks": ["LLY", "UNH", "ABT", "MDT", "DXCM", "VEEV", "HIMS", "TMO"]
    },
    "clean_energy": {
        "name": "Clean Energy",
        "color": "#276749",
        "stocks": ["NEE", "ENPH", "FSLR", "BE", "PLUG", "RUN", "ARRY", "SEDG"]
    },
    "ecommerce": {
        "name": "Consumer & E-Commerce",
        "color": "#C05621",
        "stocks": ["SHOP", "MELI", "SE", "PDD", "ETSY", "CPNG", "W", "PINS"]
    },
    "media": {
        "name": "Media & Streaming",
        "color": "#702459",
        "stocks": ["NFLX", "DIS", "SPOT", "ROKU", "TTD", "MGNI", "WMG", "PARA"]
    },
    "defense": {
        "name": "Defense & Aerospace",
        "color": "#2D3748",
        "stocks": ["LMT", "RTX", "NOC", "GD", "BA", "RKLB", "ASTS", "HII"]
    },
    "industrials": {
        "name": "Industrials & Manufacturing",
        "color": "#7B5E35",
        "stocks": ["CAT", "DE", "HON", "GE", "ETN", "PH", "IR", "EMR"]
    },
    "banking": {
        "name": "Banking & Finance",
        "color": "#1A365D",
        "stocks": ["JPM", "BAC", "GS", "MS", "C", "WFC", "AXP", "BX"]
    },
}

# ── Market Data ─────────────────────────────────────────────────────────────

def fetch_stock_data(tickers: list, period: str = "6mo") -> dict:
    """Fetch OHLCV data for a list of tickers from Yahoo Finance."""
    data = {}
    valid_tickers = []
    
    for ticker in tickers:
        try:
            t = yf.Ticker(ticker)
            hist = t.history(period=period)
            if not hist.empty and len(hist) > 20:
                data[ticker] = {
                    "history": hist,
                    "info": t.info
                }
                valid_tickers.append(ticker)
        except Exception as e:
            print(f"Failed to fetch {ticker}: {e}")
    
    return data, valid_tickers


def compute_indicators(hist: pd.DataFrame) -> dict:
    """Compute RSI, MACD, Bollinger Bands, and volume signals."""
    close = hist["Close"]
    volume = hist["Volume"]

    # RSI(14)
    delta = close.diff()
    gain = delta.clip(lower=0).rolling(14).mean()
    loss = (-delta.clip(upper=0)).rolling(14).mean()
    rs = gain / loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))

    # MACD
    ema12 = close.ewm(span=12).mean()
    ema26 = close.ewm(span=26).mean()
    macd = ema12 - ema26
    signal_line = macd.ewm(span=9).mean()
    macd_hist = macd - signal_line

    # Bollinger Bands (20, 2)
    ma20 = close.rolling(20).mean()
    std20 = close.rolling(20).std()
    bb_upper = ma20 + 2 * std20
    bb_lower = ma20 - 2 * std20
    bb_pct = (close - bb_lower) / (bb_upper - bb_lower)  # 0=lower band, 1=upper band

    # Moving averages
    ma50 = close.rolling(50).mean()
    ma20_val = ma20.iloc[-1] if len(ma20.dropna()) > 0 else close.iloc[-1]
    ma50_val = ma50.iloc[-1] if len(ma50.dropna()) > 0 else close.iloc[-1]

    # Volume ratio
    vol_avg = volume.rolling(20).mean()
    vol_ratio = volume / vol_avg

    latest = {
        "close": float(close.iloc[-1]),
        "rsi": float(rsi.iloc[-1]) if not np.isnan(rsi.iloc[-1]) else 50.0,
        "macd": float(macd.iloc[-1]) if not np.isnan(macd.iloc[-1]) else 0.0,
        "macd_signal": float(signal_line.iloc[-1]) if not np.isnan(signal_line.iloc[-1]) else 0.0,
        "macd_hist": float(macd_hist.iloc[-1]) if not np.isnan(macd_hist.iloc[-1]) else 0.0,
        "bb_pct": float(bb_pct.iloc[-1]) if not np.isnan(bb_pct.iloc[-1]) else 0.5,
        "bb_upper": float(bb_upper.iloc[-1]) if not np.isnan(bb_upper.iloc[-1]) else float(close.iloc[-1]),
        "bb_lower": float(bb_lower.iloc[-1]) if not np.isnan(bb_lower.iloc[-1]) else float(close.iloc[-1]),
        "ma20": float(ma20_val),
        "ma50": float(ma50_val),
        "vol_ratio": float(vol_ratio.iloc[-1]) if not np.isnan(vol_ratio.iloc[-1]) else 1.0,
        "price_history": close.tolist(),
        "volume_history": volume.tolist(),
        "dates": [str(d.date()) if hasattr(d, 'date') else str(d) for d in hist.index],
    }
    return latest


# ── Strategy Signal Generators ───────────────────────────────────────────────

def signal_momentum(ind: dict) -> dict:
    """RSI + MACD momentum strategy."""
    score = 0
    reasons = []
    rsi = ind["rsi"]
    if rsi > 55:
        score += 2
        reasons.append(f"RSI bullish ({rsi:.0f})")
    elif rsi < 45:
        score -= 2
        reasons.append(f"RSI bearish ({rsi:.0f})")
    if ind["macd"] > ind["macd_signal"]:
        score += 2
        reasons.append("MACD above signal")
    elif ind["macd"] < ind["macd_signal"]:
        score -= 2
        reasons.append("MACD below signal")
    if ind["macd_hist"] > 0:
        score += 1
        reasons.append("MACD histogram +")
    return {"score": score, "reasons": reasons, "max": 5}


def signal_mean_reversion(ind: dict) -> dict:
    """Bollinger Band mean reversion."""
    score = 0
    reasons = []
    bb = ind["bb_pct"]
    rsi = ind["rsi"]
    if bb < 0.1:
        score += 3
        reasons.append(f"BB lower band touch ({bb:.2f})")
    elif bb > 0.9:
        score -= 3
        reasons.append(f"BB upper band touch ({bb:.2f})")
    if rsi < 30:
        score += 2
        reasons.append(f"RSI oversold ({rsi:.0f})")
    elif rsi > 70:
        score -= 2
        reasons.append(f"RSI overbought ({rsi:.0f})")
    return {"score": score, "reasons": reasons, "max": 5}


def signal_breakout(ind: dict) -> dict:
    """Volume breakout detection."""
    score = 0
    reasons = []
    if ind["vol_ratio"] > 2.0:
        score += 3
        reasons.append(f"Volume surge {ind['vol_ratio']:.1f}x avg")
    elif ind["vol_ratio"] > 1.5:
        score += 1
        reasons.append(f"Volume elevated {ind['vol_ratio']:.1f}x avg")
    close = ind["close"]
    if close > ind["ma20"] * 1.02:
        score += 2
        reasons.append("Price above MA20 breakout")
    return {"score": score, "reasons": reasons, "max": 5}


def signal_trend_follow(ind: dict) -> dict:
    """MA crossover trend following."""
    score = 0
    reasons = []
    if ind["ma20"] > ind["ma50"]:
        score += 3
        reasons.append("Golden cross (MA20 > MA50)")
    else:
        score -= 3
        reasons.append("Death cross (MA20 < MA50)")
    if ind["close"] > ind["ma20"]:
        score += 2
        reasons.append("Price above MA20")
    else:
        score -= 2
        reasons.append("Price below MA20")
    return {"score": score, "reasons": reasons, "max": 5}


def signal_ai_hybrid(ind: dict) -> dict:
    """Combined multi-signal AI hybrid."""
    m = signal_momentum(ind)
    mr = signal_mean_reversion(ind)
    b = signal_breakout(ind)
    tf = signal_trend_follow(ind)
    combined = m["score"] + mr["score"] * 0.5 + b["score"] * 0.7 + tf["score"] * 0.8
    reasons = m["reasons"] + b["reasons"][:1] + tf["reasons"][:1]
    return {"score": combined, "reasons": reasons, "max": 5}


def signal_sector_rotation(ind: dict, sector_strength: float) -> dict:
    """Sector momentum rotation."""
    score = sector_strength * 5
    reasons = [f"Sector strength score {sector_strength:.2f}"]
    if ind["close"] > ind["ma20"]:
        score += 1
        reasons.append("Above MA20")
    return {"score": score, "reasons": reasons, "max": 6}


def signal_sentiment(ind: dict) -> dict:
    """Sentiment + technical combo. Uses cached AI sentiment when available."""
    ticker = ind.get("ticker")
    cached = sentiment_cache.get(ticker) if ticker else None

    if cached:
        sentiment_score = cached["score"]
        sentiment_label = cached.get("reasoning", "")[:60]
    else:
        sentiment_score = np.random.normal(0.55, 0.2)  # fallback until news fetched
        sentiment_label = f"random ({sentiment_score:.2f})"

    score = 0
    reasons = []
    if sentiment_score > 0.65:
        score += 2
        reasons.append(f"AI: Bullish sentiment ({sentiment_score:.2f}) — {sentiment_label}")
    elif sentiment_score < 0.35:
        score -= 2
        reasons.append(f"AI: Bearish sentiment ({sentiment_score:.2f}) — {sentiment_label}")
    else:
        reasons.append(f"AI: Neutral sentiment ({sentiment_score:.2f})")

    tech = signal_momentum(ind)
    score += tech["score"] * 0.5
    reasons += tech["reasons"][:1]
    return {"score": score, "reasons": reasons, "max": 4}


STRATEGY_SIGNALS = {
    "momentum": signal_momentum,
    "mean_rev": signal_mean_reversion,
    "breakout": signal_breakout,
    "trend_follow": signal_trend_follow,
    "ai_hybrid": signal_ai_hybrid,
    "sentiment": signal_sentiment,
}


# ── AI Stock Picker ──────────────────────────────────────────────────────────

def pick_stocks(sector_ids: list, lookback_months: int, horizon_months: int,
                target_pct: float, n_stocks: int) -> dict:
    """
    Multi-factor AI stock screener.

    Scoring model:
      40% — Price momentum (1m/3m/6m/12m weighted returns)
      30% — Technical quality (RSI position, MACD, MA trend, volume)
      30% — Risk-adjusted performance (Sharpe ratio, trend consistency R²)

    Predicted return for the forward horizon is derived from recent momentum
    extrapolated to the horizon, then blended with the Sharpe-based expected
    annual return, and adjusted by the technical quality multiplier.
    """
    # Collect deduplicated tickers and their sector labels
    ticker_sector: dict[str, str] = {}
    for sid in sector_ids:
        sec = ALL_SECTORS.get(sid)
        if not sec:
            continue
        for t in sec["stocks"]:
            if t not in ticker_sector:
                ticker_sector[t] = sec["name"]

    if not ticker_sector:
        return {"error": "No sectors selected"}

    tickers = list(ticker_sector.keys())
    period  = f"{lookback_months}mo"
    stock_data, valid = fetch_stock_data(tickers, period=period)
    if not valid:
        return {"error": "Failed to fetch stock data — try again"}

    results = []

    for ticker in valid:
        hist  = stock_data[ticker]["history"]
        info  = stock_data[ticker].get("info", {})
        close = hist["Close"].dropna()
        n     = len(close)
        if n < 30:
            continue

        price_now = float(close.iloc[-1])

        # ── Momentum returns ─────────────────────────────────────────────────
        def period_return(days: int) -> float:
            idx = max(0, n - days - 1)
            p0  = float(close.iloc[idx])
            return (price_now - p0) / p0 * 100 if p0 > 0 else 0.0

        r1m  = period_return(21)
        r3m  = period_return(63)
        r6m  = period_return(126)
        r12m = period_return(min(252, n - 2))

        # Jegadeesh-Titman: skip most recent month to reduce short-term reversal noise
        r_jt = period_return(252) - r1m if n > 253 else r12m - r1m

        mom_score = r1m * 0.20 + r3m * 0.35 + r6m * 0.25 + r_jt * 0.20

        # ── Technical quality ────────────────────────────────────────────────
        ind = compute_indicators(hist)
        rsi = ind["rsi"]

        # RSI: sweet-spot 45-65 for momentum continuation entries
        if 45 <= rsi <= 65:
            rsi_mult = 1.20
        elif 65 < rsi <= 75:
            rsi_mult = 1.00
        elif rsi > 75:
            rsi_mult = 0.75   # overbought — discount predicted gain
        elif 35 <= rsi < 45:
            rsi_mult = 1.10   # mildly oversold — mean reversion opportunity
        else:
            rsi_mult = 0.85   # deeply oversold — uncertain

        macd_bull  = 1 if ind["macd"] > ind["macd_signal"] and ind["macd_hist"] > 0 else 0
        trend_bull = 1 if ind["ma20"] > ind["ma50"] else 0
        above_ma20 = 1 if price_now > ind["ma20"] else 0
        vol_conf   = 1 if ind["vol_ratio"] > 1.2 else 0
        bb_ok      = 1 if 0.15 < ind["bb_pct"] < 0.85 else 0   # not at extremes

        tech_signals  = macd_bull + trend_bull + above_ma20 + vol_conf + bb_ok
        tech_score_01 = tech_signals / 5.0    # 0.0 – 1.0
        tech_mult     = 0.80 + tech_score_01 * 0.40   # 0.80 – 1.20

        # ── Risk-adjusted performance ────────────────────────────────────────
        daily_rets = close.pct_change().dropna()
        if len(daily_rets) > 20 and daily_rets.std() > 0:
            sharpe = float(daily_rets.mean() / daily_rets.std()) * (252 ** 0.5)
        else:
            sharpe = 0.0

        # Trend consistency via linear-regression R²
        x     = np.arange(min(n, 63), dtype=float)   # last 3 months
        y     = close.values[-len(x):]
        if len(x) > 5:
            x_m  = x.mean(); y_m = y.mean()
            ss_tot = float(((y - y_m) ** 2).sum())
            slope  = float(((x - x_m) * (y - y_m)).sum() / (((x - x_m) ** 2).sum() + 1e-9))
            y_hat  = y_m + slope * (x - x_m)
            ss_res = float(((y - y_hat) ** 2).sum())
            r2     = max(0.0, 1.0 - ss_res / (ss_tot + 1e-9))
        else:
            r2 = 0.0

        # ── Composite score (for ranking) ────────────────────────────────────
        composite = (
            mom_score          * 0.40
            + tech_score_01    * 40 * 0.30
            + sharpe           * 8  * 0.20
            + r2               * 10 * 0.10
        )

        # ── Predicted return for forward horizon ─────────────────────────────
        # Base: momentum extrapolated to horizon
        if horizon_months <= 3:
            base_pred = r3m * (horizon_months / 3.0)
        elif horizon_months <= 6:
            base_pred = r3m * 0.5 + r6m * (horizon_months / 6.0) * 0.5
        else:
            base_pred = r6m * 0.5 + r12m * (horizon_months / 12.0) * 0.5

        # Sharpe-based expected annual ~ sharpe * 15%; scale to horizon
        sharpe_pred = sharpe * 15.0 * (horizon_months / 12.0)

        # Blend momentum and sharpe-based, then apply technical multiplier
        predicted = (base_pred * 0.60 + sharpe_pred * 0.40) * tech_mult * rsi_mult

        # ── Confidence (0-100) ───────────────────────────────────────────────
        aligned = sum([
            r1m  > 0,
            r3m  > 0,
            r6m  > 0,
            bool(macd_bull),
            bool(trend_bull),
            bool(vol_conf),
            45 <= rsi <= 75,
            sharpe > 0.5,
            r2 > 0.5,
        ])
        confidence = int(aligned / 9 * 100)

        results.append({
            "ticker":           ticker,
            "name":             info.get("shortName", ticker),
            "sector":           ticker_sector.get(ticker, ""),
            "price":            round(price_now, 2),
            "r1m":              round(r1m,  2),
            "r3m":              round(r3m,  2),
            "r6m":              round(r6m,  2),
            "r12m":             round(r12m, 2),
            "rsi":              round(rsi,  1),
            "sharpe":           round(sharpe, 2),
            "r2":               round(r2, 2),
            "macd_bull":        bool(macd_bull),
            "trend_bull":       bool(trend_bull),
            "vol_conf":         bool(vol_conf),
            "composite_score":  round(composite, 2),
            "predicted_return": round(predicted, 2),
            "confidence":       confidence,
        })

    # Filter by minimum target, sort by composite score
    qualified = sorted(
        [r for r in results if r["predicted_return"] >= target_pct],
        key=lambda x: -x["composite_score"]
    )

    return {
        "picks":            qualified[:n_stocks],
        "total_screened":   len(results),
        "total_qualified":  len(qualified),
        "horizon_months":   horizon_months,
        "lookback_months":  lookback_months,
        "target_pct":       target_pct,
    }


# ── Simulation Engine ────────────────────────────────────────────────────────

def run_simulation(
    capital: float,
    strategy: str,
    months: int,
    risk: str,
    sector_ids: list,
    stock_data: dict
) -> dict:
    """
    Core simulation engine.
    Uses real historical price paths from Yahoo Finance.
    Replays prices day by day and applies strategy signals.
    """
    stop_loss = {"low": 0.05, "med": 0.10, "high": 0.20}[risk]
    take_profit = stop_loss * 2.5
    trade_freq = {
        "momentum": 12, "mean_rev": 9, "ai_hybrid": 11,
        "breakout": 7, "pairs": 13, "sector_rot": 6,
        "trend_follow": 5, "sentiment": 9
    }.get(strategy, 9)

    # Collect tickers for selected sectors
    tickers = []
    for sid in sector_ids:
        if sid in SECTORS:
            tickers.extend(SECTORS[sid]["stocks"])

    # Filter to tickers we actually have data for
    available = [t for t in tickers if t in stock_data]
    if not available:
        return {"error": "No valid stock data fetched"}

    # Build price series: align all tickers to common dates
    price_series = {}
    for t in available:
        hist = stock_data[t]["history"]
        price_series[t] = hist["Close"].values.tolist()

    min_len = min(len(v) for v in price_series.values())
    trading_days = min(months * 21, min_len - 1)

    positions = {}
    cash = capital
    trades = []
    equity_curve = [capital]
    per_stock_alloc = capital * 0.85 / len(available)

    # Open initial positions
    for t in available[:min(8, len(available))]:
        price = price_series[t][0]
        shares = int(per_stock_alloc / price)
        if shares > 0 and cash >= shares * price:
            positions[t] = {
                "ticker": t,
                "name": stock_data[t].get("info", {}).get("shortName", t),
                "sector": next((SECTORS[s]["name"] for s in sector_ids if t in SECTORS.get(s, {}).get("stocks", [])), "AI"),
                "shares": shares,
                "entry": price,
                "current": price,
                "high_water": price,
                "day_opened": 0
            }
            cash -= shares * price

    # Replay price history day by day
    for day in range(1, trading_days + 1):
        # Update prices and check stops
        for t in list(positions.keys()):
            if day < len(price_series[t]):
                new_price = price_series[t][day]
                positions[t]["current"] = new_price
                p = positions[t]

                if new_price > p["high_water"]:
                    p["high_water"] = new_price

                pct_change = (new_price - p["entry"]) / p["entry"]
                hit_stop = pct_change <= -stop_loss
                hit_tp = pct_change >= take_profit

                if hit_stop or hit_tp:
                    pnl = (new_price - p["entry"]) * p["shares"]
                    trades.append({
                        "day": day,
                        "action": "SELL",
                        "ticker": t,
                        "name": p["name"],
                        "sector": p["sector"],
                        "price": round(new_price, 2),
                        "shares": p["shares"],
                        "pnl": round(pnl, 2),
                        "signal": "Take profit" if hit_tp else "Stop-loss hit",
                        "strategy": strategy
                    })
                    cash += new_price * p["shares"]
                    del positions[t]

        # Generate trade signals
        if day % max(1, round(21 / trade_freq)) == 0:
            # Score available stocks
            scored = []
            for t in available:
                if t not in positions and day < len(price_series[t]) - 5:
                    window = min(60, day)
                    if window < 20:
                        continue
                    prices = price_series[t][max(0, day-window):day]
                    volumes = stock_data[t]["history"]["Volume"].values.tolist()
                    vols_window = volumes[max(0, day-window):day]

                    hist_slice = pd.DataFrame({
                        "Close": prices,
                        "Volume": vols_window[:len(prices)]
                    })
                    ind = compute_indicators(hist_slice)
                    ind["ticker"] = t

                    if strategy in STRATEGY_SIGNALS:
                        sig = STRATEGY_SIGNALS[strategy](ind)
                    else:
                        sig = signal_ai_hybrid(ind)

                    if sig["score"] > 1.5:
                        scored.append((t, sig, price_series[t][day]))

            scored.sort(key=lambda x: -x[1]["score"])

            # Buy top candidates
            for t, sig, price in scored[:2]:
                if cash < 500:
                    break
                shares = int(min(cash * 0.4, per_stock_alloc * 1.2) / price)
                if shares > 0:
                    positions[t] = {
                        "ticker": t,
                        "name": stock_data[t].get("info", {}).get("shortName", t),
                        "sector": next((SECTORS[s]["name"] for s in sector_ids if t in SECTORS.get(s, {}).get("stocks", [])), "AI"),
                        "shares": shares,
                        "entry": price,
                        "current": price,
                        "high_water": price,
                        "day_opened": day
                    }
                    cash -= shares * price
                    trades.append({
                        "day": day,
                        "action": "BUY",
                        "ticker": t,
                        "name": positions[t]["name"],
                        "sector": positions[t]["sector"],
                        "price": round(price, 2),
                        "shares": shares,
                        "pnl": 0,
                        "signal": "; ".join(sig["reasons"][:2]),
                        "strategy": strategy
                    })

        # Equity snapshot
        port_val = cash + sum(p["shares"] * p["current"] for p in positions.values())
        equity_curve.append(port_val)

    # Final portfolio value
    final_val = cash + sum(p["shares"] * p["current"] for p in positions.values())
    ret = (final_val - capital) / capital

    # Performance metrics
    sells = [t for t in trades if t["action"] == "SELL"]
    wins = [t for t in sells if t["pnl"] > 0]
    daily_rets = [(equity_curve[i] - equity_curve[i-1]) / equity_curve[i-1]
                  for i in range(1, len(equity_curve)) if equity_curve[i-1] > 0]
    avg_r = np.mean(daily_rets) if daily_rets else 0
    std_r = np.std(daily_rets) if daily_rets else 1e-9
    sharpe = (avg_r / std_r) * np.sqrt(252) if std_r > 0 else 0
    peak = capital
    max_dd = 0
    for v in equity_curve:
        if v > peak:
            peak = v
        dd = (peak - v) / peak
        if dd > max_dd:
            max_dd = dd

    # Indicator snapshots for signal heatmap
    signal_snapshots = {}
    for t in list(positions.keys())[:10]:
        window = min(60, trading_days)
        prices = price_series[t][max(0, trading_days-window):trading_days]
        volumes = stock_data[t]["history"]["Volume"].values.tolist()
        vols_w = volumes[max(0, trading_days-window):trading_days]
        hist_slice = pd.DataFrame({"Close": prices, "Volume": vols_w[:len(prices)]})
        ind = compute_indicators(hist_slice)
        signal_snapshots[t] = {
            "rsi": round(ind["rsi"], 1),
            "macd_hist": round(ind["macd_hist"], 4),
            "bb_pct": round(ind["bb_pct"] * 100, 1),
            "vol_ratio": round(ind["vol_ratio"], 2),
            "ma_cross": round((ind["ma20"] - ind["ma50"]) / ind["ma50"] * 100, 2) if ind["ma50"] > 0 else 0,
        }

    return {
        "final_val": round(final_val, 2),
        "capital": capital,
        "return": round(ret * 100, 2),
        "return_dollar": round(final_val - capital, 2),
        "sharpe": round(sharpe, 2),
        "win_rate": round(len(wins) / max(1, len(sells)) * 100, 1),
        "max_drawdown": round(max_dd * 100, 2),
        "total_trades": len(trades),
        "closed_trades": len(sells),
        "winning_trades": len(wins),
        "trades": trades,
        "positions": list(positions.values()),
        "equity_curve": [round(v, 2) for v in equity_curve],
        "daily_pnl": [round(equity_curve[i] - equity_curve[i-1], 2) for i in range(1, len(equity_curve))],
        "signal_snapshots": signal_snapshots,
        "trading_days": trading_days,
    }


# ── News & AI Research ──────────────────────────────────────────────────────

YAHOO_RSS_URL = "https://feeds.finance.yahoo.com/rss/2.0/headline?s={ticker}&region=US&lang=en-US"
RSS_NAMESPACE = "{http://www.w3.org/2005/Atom}"


def fetch_yahoo_rss(tickers: list, max_per_ticker: int = 5) -> dict:
    """Fetch latest news headlines from Yahoo Finance RSS for each ticker."""
    headlines = {}
    ssl_ctx = ssl.create_default_context(cafile=certifi.where())
    for ticker in tickers:
        try:
            url = YAHOO_RSS_URL.format(ticker=ticker)
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=8, context=ssl_ctx) as resp:
                xml_data = resp.read()
            root = ET.fromstring(xml_data)
            channel = root.find("channel")
            if channel is None:
                continue
            items = channel.findall("item")[:max_per_ticker]
            ticker_headlines = []
            for item in items:
                title_el = item.find("title")
                pub_el = item.find("pubDate")
                if title_el is not None and title_el.text:
                    ticker_headlines.append({
                        "title": title_el.text.strip(),
                        "pubDate": pub_el.text.strip() if pub_el is not None else ""
                    })
            if ticker_headlines:
                headlines[ticker] = ticker_headlines
        except Exception as e:
            print(f"RSS fetch failed for {ticker}: {e}")
    return headlines


def analyze_with_openai(headlines_by_ticker: dict, sector_ids: list) -> dict:
    """Send headlines to OpenAI GPT and get structured sentiment analysis."""
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise ValueError("OPENAI_API_KEY environment variable is not set")

    try:
        from openai import OpenAI
    except ImportError:
        raise ImportError("openai package not installed — run: pip install openai>=1.0.0")

    client = OpenAI(api_key=api_key)

    sector_names = [SECTORS[s]["name"] for s in sector_ids if s in SECTORS]

    # Build compact headline text
    headline_lines = []
    for ticker, items in headlines_by_ticker.items():
        for h in items:
            headline_lines.append(f"{ticker}: {h['title']}")

    headline_text = "\n".join(headline_lines) if headline_lines else "No headlines available."

    prompt = f"""You are a financial analyst specializing in AI and technology stocks.

Analyze the following recent news headlines for stocks in these sectors: {', '.join(sector_names)}.

Headlines:
{headline_text}

Return ONLY a valid JSON object with this exact structure:
{{
  "tickers": [
    {{"ticker": "NVDA", "score": 0.75, "reasoning": "brief one-sentence reason"}}
  ],
  "sector_summaries": [
    {{"sector": "Semiconductors", "outlook": "bullish", "summary": "one or two sentence summary"}}
  ],
  "market_themes": ["theme 1", "theme 2", "theme 3"]
}}

Rules:
- score is a float from 0.0 (very bearish) to 1.0 (very bullish), 0.5 = neutral
- Only include tickers that appear in the headlines
- outlook must be exactly: "bullish", "neutral", or "bearish"
- Return 2-4 market themes
- No markdown, no extra keys, no explanation outside the JSON"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.3,
        max_tokens=1200,
    )

    raw = response.choices[0].message.content.strip()
    # Strip markdown code fences if present
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[-1]
        if raw.endswith("```"):
            raw = raw[:-3]

    analysis = json.loads(raw)

    # Validate minimal structure
    if "tickers" not in analysis or "sector_summaries" not in analysis:
        raise ValueError("Unexpected response structure from OpenAI")

    return analysis


@app.route("/api/all-sectors")
def api_all_sectors():
    """Return the full sector universe used by the Stock Picker."""
    return jsonify({
        sid: {"name": s["name"], "color": s["color"], "stocks": s["stocks"]}
        for sid, s in ALL_SECTORS.items()
    })


@app.route("/api/pick-stocks", methods=["POST"])
def api_pick_stocks():
    """Run the AI Stock Picker for the requested sectors and parameters."""
    body         = request.json or {}
    sector_ids   = body.get("sectors",  list(ALL_SECTORS.keys()))
    lookback     = int(body.get("lookback",  12))
    horizon      = int(body.get("horizon",    3))
    target_pct   = float(body.get("target_pct", 10.0))
    n_stocks     = int(body.get("n_stocks",   10))

    # Clamp to reasonable bounds
    lookback  = max(3,  min(lookback,  24))
    horizon   = max(1,  min(horizon,   24))
    target_pct = max(0.0, min(target_pct, 500.0))
    n_stocks  = max(1,  min(n_stocks,  50))

    try:
        result = pick_stocks(sector_ids, lookback, horizon, target_pct, n_stocks)
        if "error" in result:
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/news", methods=["POST"])
def api_news():
    """Fetch Yahoo Finance RSS headlines and analyze with OpenAI."""
    body = request.json or {}
    sector_ids = body.get("sectors", list(SECTORS.keys()))

    # Collect tickers from requested sectors
    tickers = []
    for sid in sector_ids:
        if sid in SECTORS:
            tickers.extend(SECTORS[sid]["stocks"])
    tickers = list(dict.fromkeys(tickers))  # deduplicate

    # 1. Fetch RSS headlines
    try:
        headlines = fetch_yahoo_rss(tickers)
    except Exception as e:
        return jsonify({"error": f"RSS fetch failed: {str(e)}"}), 500

    if not headlines:
        return jsonify({"error": "No headlines found for the selected sectors"}), 404

    # 2. Analyze with OpenAI
    try:
        analysis = analyze_with_openai(headlines, sector_ids)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"AI analysis failed: {str(e)}"}), 500

    # 3. Update sentiment cache
    now_str = datetime.now().isoformat()
    for item in analysis.get("tickers", []):
        ticker = item.get("ticker")
        if ticker:
            sentiment_cache[ticker] = {
                "score": float(item.get("score", 0.5)),
                "reasoning": item.get("reasoning", ""),
                "updated_at": now_str,
            }

    return jsonify({
        "news": headlines,
        "analysis": analysis,
        "cached_at": now_str,
        "tickers_analyzed": len(analysis.get("tickers", [])),
        "headlines_fetched": sum(len(v) for v in headlines.values()),
    })


# ── Excel Export ─────────────────────────────────────────────────────────────

def generate_excel(result: dict, strategy: str, sector_ids: list) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", start_color="0C447C")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    body_font = Font(size=10)
    buy_fill = PatternFill("solid", start_color="E1F5EE")
    sell_fill = PatternFill("solid", start_color="FCEBEB")
    border = Border(bottom=Side(style="thin", color="DDDDDD"))

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.fill = hdr_fill
            cell.font = hdr_font

    def set_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "AI Sector Paper Trading — Simulation Report"
    ws["A1"].font = Font(bold=True, size=14, color="0C447C")
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}   Strategy: {strategy}   Capital: ${result['capital']:,.0f}"
    ws["A2"].font = Font(size=10, color="888888")
    ws.merge_cells("A2:D2")
    ws.append([])
    ws.append(["Metric", "Value"])
    style_header(ws, 4)
    for row in [
        ("Starting capital", f"${result['capital']:,.2f}"),
        ("Final portfolio value", f"${result['final_val']:,.2f}"),
        ("Total return", f"{result['return']:+.2f}%"),
        ("Total P&L", f"${result['return_dollar']:+,.2f}"),
        ("Sharpe ratio", result["sharpe"]),
        ("Win rate", f"{result['win_rate']:.1f}%"),
        ("Max drawdown", f"-{result['max_drawdown']:.2f}%"),
        ("Total trades", result["total_trades"]),
        ("Closed trades", result["closed_trades"]),
        ("Winning trades", result["winning_trades"]),
    ]:
        ws.append(list(row))
        for c in ws[ws.max_row]:
            c.font = body_font
            c.border = border
    set_widths(ws, [28, 20])

    # Sheet 2: Trade Log
    ws2 = wb.create_sheet("Trade Log")
    ws2.append(["Day", "Action", "Ticker", "Company", "Sector", "Price ($)", "Shares", "P&L ($)", "Signal", "Strategy"])
    style_header(ws2)
    for t in result["trades"]:
        row = [t["day"], t["action"], t["ticker"], t["name"], t["sector"],
               t["price"], t["shares"],
               t["pnl"] if t["action"] == "SELL" else "",
               t["signal"], t["strategy"]]
        ws2.append(row)
        fill = sell_fill if t["action"] == "SELL" else buy_fill
        for c in ws2[ws2.max_row]:
            c.fill = fill
            c.font = body_font
    set_widths(ws2, [6, 8, 8, 20, 22, 12, 8, 12, 32, 16])

    # Sheet 3: Equity Curve
    ws3 = wb.create_sheet("Equity Curve")
    ws3.append(["Day", "Portfolio Value ($)", "Daily P&L ($)", "Cumulative Return (%)"])
    style_header(ws3)
    capital = result["capital"]
    for i, v in enumerate(result["equity_curve"]):
        dpnl = result["daily_pnl"][i-1] if i > 0 else 0
        ret_pct = (v - capital) / capital * 100
        ws3.append([i, round(v, 2), round(dpnl, 2), round(ret_pct, 4)])
        for c in ws3[ws3.max_row]:
            c.font = body_font
    set_widths(ws3, [8, 22, 16, 22])

    # Sheet 4: Open Positions
    ws4 = wb.create_sheet("Open Positions")
    ws4.append(["Ticker", "Company", "Sector", "Shares", "Entry ($)", "Current ($)", "Unrealized P&L ($)", "Return (%)"])
    style_header(ws4)
    for p in result["positions"]:
        pnl = (p["current"] - p["entry"]) * p["shares"]
        ret_pct = (p["current"] - p["entry"]) / p["entry"] * 100 if p["entry"] > 0 else 0
        ws4.append([p["ticker"], p.get("name", ""), p.get("sector", ""),
                    p["shares"], round(p["entry"], 2), round(p["current"], 2),
                    round(pnl, 2), round(ret_pct, 2)])
        for c in ws4[ws4.max_row]:
            c.font = body_font
    set_widths(ws4, [8, 20, 22, 8, 12, 12, 20, 12])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Autonomous Trading Helpers ────────────────────────────────────────────────

def pick_risk_level_from_market() -> str:
    """
    Determine risk level dynamically from SPY 20-day annualised volatility.
      vol > 28%  -> low  (5% stop-loss / 12.5% take-profit)
      vol > 16%  -> med  (10% stop-loss / 25% take-profit)
      vol <= 16% -> high (20% stop-loss / 50% take-profit)
    """
    try:
        spy = yf.Ticker("SPY")
        hist = spy.history(period="2mo")
        rets = hist["Close"].pct_change().dropna()
        vol  = float(rets.rolling(20).std().iloc[-1]) * np.sqrt(252)
        if vol > 0.28:
            level = "low"
        elif vol > 0.16:
            level = "med"
        else:
            level = "high"
        print(f"  Market volatility (annualised) {vol*100:.1f}%  -> risk level: {level}")
        return level
    except Exception as e:
        print(f"  Risk level fetch failed, defaulting to med: {e}")
        return "med"


def best_signal_for(ind: dict) -> tuple[str, dict]:
    """
    Run all strategies against a single ticker and return (best_strategy_name, best_signal).
    Scores are normalised to [0, 10] against each strategy's max so they are comparable.
    """
    best_name  = "ai_hybrid"
    best_score = -999
    best_sig   = {"score": 0, "reasons": [], "max": 5}

    for name, fn in STRATEGY_SIGNALS.items():
        try:
            sig  = fn(ind)
            smax = sig.get("max", 5) or 5
            normalized = sig["score"] / smax * 10   # 归一化到 0-10
        except Exception:
            continue

        if normalized > best_score:
            best_score = normalized
            best_name  = name
            best_sig   = sig
            best_sig["normalized"] = normalized  # carry normalised score

    return best_name, best_sig


# ── Auto-Trader ───────────────────────────────────────────────────────────────

def run_auto_trade():
    """
    Autonomous daily paper-trading engine (no human restrictions):
    1. Determine risk level automatically from SPY market volatility
    2. Scan all 6 AI sectors (~48 US stocks) in the universe
    3. Run every strategy on every stock; pick the highest-scoring one
    4. Refresh news sentiment via OpenAI and layer it on top of scores
    5. Buy the strongest candidates within position and cash limits
    6. Manage existing positions: stop-loss / take-profit / trailing stop
    Goal: maximise long-term portfolio value above the initial capital
    """
    today = datetime.now().isoformat()[:10]

    print(f"\n{'='*60}")
    print(f"  Auto-paper-trade (Autonomous): {today}")
    print(f"{'='*60}")

    # 1. Load portfolio state
    portfolio = load_portfolio()
    initial   = portfolio["initial"]
    cash      = portfolio["cash"]
    positions = portfolio["positions"]
    print(f"  Cash: ${cash:,.2f}  Positions: {len(positions)}  Initial: ${initial:,.2f}")

    # 2. Determine risk level autonomously
    risk        = pick_risk_level_from_market()
    stop_loss   = {"low": 0.05, "med": 0.10, "high": 0.20}[risk]
    take_profit = stop_loss * 2.5

    # 3. Build full stock universe — all 6 sectors
    all_sector_ids = list(SECTORS.keys())
    universe = []
    for sid in all_sector_ids:
        universe.extend(SECTORS[sid]["stocks"])
    universe = list(dict.fromkeys(universe))  # deduplicate, preserve order

    # Include current holdings so we always refresh their prices
    all_tickers = list(dict.fromkeys(list(positions.keys()) + universe))
    print(f"  Universe: {len(all_tickers)} stocks (all sectors)")

    # 4. Fetch market data (3 months for indicator calculation)
    stock_data, valid = fetch_stock_data(all_tickers, period="3mo")
    if not valid:
        print("  ERROR: Failed to fetch any stock data — skipping run")
        return
    print(f"  Fetched data: {len(valid)} stocks OK")

    # 5. Refresh news sentiment for all sectors (requires OPENAI_API_KEY)
    if os.environ.get("OPENAI_API_KEY"):
        try:
            headlines = fetch_yahoo_rss(valid[:30], max_per_ticker=3)
            if headlines:
                analysis = analyze_with_openai(headlines, all_sector_ids)
                now_str = datetime.now().isoformat()
                for item in analysis.get("tickers", []):
                    t = item.get("ticker")
                    if t:
                        sentiment_cache[t] = {
                            "score":      float(item.get("score", 0.5)),
                            "reasoning":  item.get("reasoning", ""),
                            "updated_at": now_str,
                        }
                print(f"  Sentiment updated: {len(sentiment_cache)} tickers")
        except Exception as e:
            print(f"  News fetch warning (trading continues): {e}")

    trades_today = []

    # 6. Manage existing positions: stop-loss / take-profit / trailing stop
    for ticker in list(positions.keys()):
        if ticker not in stock_data:
            print(f"  {ticker} no data — holding position")
            continue
        hist  = stock_data[ticker]["history"]
        price = float(hist["Close"].iloc[-1])
        pos   = positions[ticker]
        pos["current"] = price
        if price > pos.get("high_water", pos["entry"]):
            pos["high_water"] = price

        pct      = (price - pos["entry"]) / pos["entry"]
        hit_stop = pct <= -stop_loss
        hit_tp   = pct >= take_profit

        # Trailing stop: close if price pulls back 60% of the stop distance from high-water
        trailing_trigger = (pos["high_water"] - price) / pos["high_water"] >= stop_loss * 0.6

        if hit_stop or hit_tp or (trailing_trigger and pct > 0):
            pnl    = (price - pos["entry"]) * pos["shares"]
            cash  += price * pos["shares"]
            if hit_tp:
                reason = "Take-profit"
            elif hit_stop:
                reason = "Stop-loss"
            else:
                reason = f"Trailing stop ({pct*100:+.1f}%)"
            trades_today.append({
                "date":     today,
                "action":   "SELL",
                "ticker":   ticker,
                "name":     pos.get("name", ticker),
                "sector":   pos.get("sector", "AI"),
                "price":    round(price, 2),
                "shares":   pos["shares"],
                "pnl":      round(pnl, 2),
                "signal":   reason,
                "strategy": "autonomous",
            })
            print(f"  SELL {ticker:<6} {reason:<24} PnL ${pnl:+,.2f}  ({pct*100:+.1f}%)")
            del positions[ticker]

    # 7. Scan for buy candidates — run all strategies, pick best per ticker
    max_pos      = config.MAX_POSITIONS
    cash_reserve = initial * config.CASH_RESERVE_RATIO
    avail_cash   = cash - cash_reserve

    if len(positions) < max_pos and avail_cash > 500:
        candidates = []

        for ticker in valid:
            if ticker in positions:
                continue
            hist = stock_data[ticker]["history"]
            if len(hist) < 30:
                continue

            ind           = compute_indicators(hist)
            ind["ticker"] = ticker

            # Run all strategies, take the highest-scoring one
            strat_name, sig = best_signal_for(ind)
            norm_score      = sig.get("normalized", 0)

            # Layer in news sentiment (same scale)
            cached = sentiment_cache.get(ticker)
            if cached:
                s = cached["score"]
                if s > 0.65:
                    bonus = (s - 0.5) * 4       # up to +2 pts
                    norm_score  += bonus
                    sig["reasons"].append(f"Bullish news ({s:.2f})")
                elif s < 0.35:
                    penalty = (0.5 - s) * 6     # up to -3 pts
                    norm_score  -= penalty
                    sig["reasons"].append(f"Bearish news ({s:.2f})")

            # Only consider clearly bullish signals
            if norm_score >= config.SIGNAL_THRESHOLD:
                price = float(hist["Close"].iloc[-1])
                sector_name = next(
                    (SECTORS[sid]["name"] for sid in all_sector_ids
                     if ticker in SECTORS.get(sid, {}).get("stocks", [])),
                    "US Market"
                )
                candidates.append((ticker, sig, price, strat_name, norm_score, sector_name))

        # Sort by composite score descending
        candidates.sort(key=lambda x: -x[4])

        if candidates:
            print("  Buy candidates (top 5): " +
                  ", ".join(f"{c[0]}({c[3]},{c[4]:.1f})" for c in candidates[:5]))

        # Enter up to 3 positions per run (staged entry)
        for ticker, sig, price, strat_name, norm_score, sector_name in candidates[:3]:
            slots_left = max_pos - len(positions)
            if slots_left <= 0 or avail_cash < 500:
                break

            # Score-weighted position sizing: higher score = larger allocation (15%–40%)
            weight = min(0.40, max(0.15, norm_score / 20))
            alloc  = min(avail_cash * weight,
                         avail_cash / max(slots_left, 1))
            shares = int(alloc / price)

            if shares > 0 and (cash - shares * price) >= cash_reserve:
                name = stock_data[ticker].get("info", {}).get("shortName", ticker)
                positions[ticker] = {
                    "ticker":     ticker,
                    "name":       name,
                    "sector":     sector_name,
                    "shares":     shares,
                    "entry":      price,
                    "current":    price,
                    "high_water": price,
                    "day_opened": today,
                }
                cost       = shares * price
                cash      -= cost
                avail_cash -= cost
                trades_today.append({
                    "date":     today,
                    "action":   "BUY",
                    "ticker":   ticker,
                    "name":     name,
                    "sector":   sector_name,
                    "price":    round(price, 2),
                    "shares":   shares,
                    "pnl":      0.0,
                    "signal":   "; ".join(sig["reasons"][:3]),
                    "strategy": strat_name,
                })
                print(f"  BUY  {ticker:<6} {shares:>4} shares @ ${price:>8.2f}"
                      f"  strategy={strat_name}  score={norm_score:.1f}")

    # 8. Persist portfolio and trade log
    portfolio["cash"]      = cash
    portfolio["positions"] = positions
    save_portfolio(portfolio)
    save_trades(trades_today)

    port_value = cash + sum(p["shares"] * p.get("current", p["entry"])
                            for p in positions.values())
    notes = (
        f"{len(trades_today)} trades: " +
        ", ".join(f"{t['action']} {t['ticker']}({t['strategy']})" for t in trades_today)
        if trades_today else "No trades"
    )
    save_daily_run(today, port_value, cash, initial, len(trades_today), notes)

    ret = (port_value - initial) / initial * 100
    print(f"\n  Portfolio value: ${port_value:,.2f}  Total return: {ret:+.2f}%"
          f"  risk={risk}  trades today: {len(trades_today)}")
    print(f"{'='*60}\n")


# ── Scheduler ─────────────────────────────────────────────────────────────────

def start_scheduler():
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger
    import pytz
    scheduler = BackgroundScheduler(daemon=True)
    tz = pytz.timezone(config.SCHEDULER_TIMEZONE)
    scheduler.add_job(
        run_auto_trade,
        trigger=CronTrigger(
            hour=config.SCHEDULER_HOUR,
            minute=config.SCHEDULER_MINUTE,
            timezone=tz,
        ),
        id="daily_auto_trade",
        replace_existing=True,
        misfire_grace_time=3600,
    )
    scheduler.start()
    print(f"  Scheduler started — daily auto-trade at "
          f"{config.SCHEDULER_HOUR:02d}:{config.SCHEDULER_MINUTE:02d} {config.SCHEDULER_TIMEZONE}")
    return scheduler


# ── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html", sectors=SECTORS)


@app.route("/api/sectors")
def api_sectors():
    return jsonify(SECTORS)


@app.route("/api/simulate", methods=["POST"])
def api_simulate():
    body = request.json
    capital = float(body.get("capital", 10000))
    strategy = body.get("strategy", "ai_hybrid")
    months = int(body.get("months", 3))
    risk = body.get("risk", "med")
    sector_ids = body.get("sectors", ["semis", "cloud_ai"])

    # Collect tickers
    tickers = []
    for sid in sector_ids:
        if sid in SECTORS:
            tickers.extend(SECTORS[sid]["stocks"])
    tickers = list(dict.fromkeys(tickers))  # deduplicate, preserve order

    # Fetch real data
    stock_data, valid = fetch_stock_data(tickers, period=f"{max(months*2, 3)}mo")

    if not valid:
        return jsonify({"error": "Could not fetch any stock data. Check your internet connection."}), 500

    result = run_simulation(capital, strategy, months, risk, sector_ids, stock_data)

    if "error" in result:
        return jsonify(result), 500

    return jsonify(result)


@app.route("/api/quote/<ticker>")
def api_quote(ticker):
    """Get latest quote + indicators for a single ticker."""
    try:
        t = yf.Ticker(ticker.upper())
        hist = t.history(period="3mo")
        if hist.empty:
            return jsonify({"error": "No data"}), 404
        ind = compute_indicators(hist)
        info = t.info
        return jsonify({
            "ticker": ticker.upper(),
            "name": info.get("shortName", ticker),
            "price": ind["close"],
            "rsi": round(ind["rsi"], 1),
            "macd_hist": round(ind["macd_hist"], 4),
            "bb_pct": round(ind["bb_pct"], 3),
            "vol_ratio": round(ind["vol_ratio"], 2),
            "ma20": round(ind["ma20"], 2),
            "ma50": round(ind["ma50"], 2),
            "price_history": ind["price_history"][-60:],
            "dates": ind["dates"][-60:],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export", methods=["POST"])
def api_export():
    body = request.json
    result = body.get("result")
    strategy = body.get("strategy", "ai_hybrid")
    sectors = body.get("sectors", [])

    if not result:
        return jsonify({"error": "No result data"}), 400

    xlsx_bytes = generate_excel(result, strategy, sectors)
    return send_file(
        io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"ai_trading_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )


@app.route("/api/portfolio")
def api_portfolio():
    """Return live portfolio state (auto-trade positions + daily snapshots)."""
    try:
        portfolio = load_portfolio()
        positions_list = []
        total_market = 0.0

        tickers = list(portfolio["positions"].keys())
        if tickers:
            # Use fast_info for true live/intraday price — avoids the
            # "daily close only" limitation of history(period="5d") which
            # returns yesterday's close during market hours, causing 0% return.
            live_prices = {}
            for ticker in tickers:
                try:
                    fi = yf.Ticker(ticker).fast_info
                    price = fi.get("lastPrice") or fi.get("regularMarketPrice")
                    if price and float(price) > 0:
                        live_prices[ticker] = float(price)
                except Exception:
                    pass
            # Fallback to daily history if fast_info fails for a ticker
            if len(live_prices) < len(tickers):
                missing = [t for t in tickers if t not in live_prices]
                stock_data, _ = fetch_stock_data(missing, period="5d")
                for ticker in missing:
                    if ticker in stock_data:
                        hist = stock_data[ticker]["history"]
                        live_prices[ticker] = float(hist["Close"].iloc[-1])
            # Apply prices
            for ticker, pos in portfolio["positions"].items():
                if ticker in live_prices:
                    pos["current"] = live_prices[ticker]

        for ticker, pos in portfolio["positions"].items():
            mkt_val      = pos["shares"] * pos["current"]
            total_market += mkt_val
            pnl          = (pos["current"] - pos["entry"]) * pos["shares"]
            pct          = ((pos["current"] - pos["entry"]) / pos["entry"] * 100
                            if pos["entry"] > 0 else 0)
            positions_list.append({
                "ticker":       ticker,
                "name":         pos.get("name", ticker),
                "sector":       pos.get("sector", ""),
                "shares":       pos["shares"],
                "entry":        round(pos["entry"], 2),
                "current":      round(pos["current"], 2),
                "pnl":          round(pnl, 2),
                "pct":          round(pct, 2),
                "market_value": round(mkt_val, 2),
                "date_opened":  pos.get("day_opened", ""),
            })

        total_value = portfolio["cash"] + total_market
        initial     = portfolio["initial"]
        ret         = (total_value - initial) / initial * 100 if initial > 0 else 0

        conn = get_db()
        runs = conn.execute(
            """SELECT date, portfolio_value, return_pct, trades_count, notes
               FROM daily_runs ORDER BY date ASC LIMIT 180"""
        ).fetchall()
        conn.close()

        return jsonify({
            "cash":          round(portfolio["cash"], 2),
            "initial":       initial,
            "started_at":    portfolio.get("started_at", ""),
            "total_value":   round(total_value, 2),
            "return_pct":    round(ret, 2),
            "return_dollar": round(total_value - initial, 2),
            "positions":     positions_list,
            "daily_runs":    [dict(r) for r in runs],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/history")
def api_history():
    """返回全量交易记录（自动交易产生的所有买卖）。"""
    try:
        conn = get_db()
        trades = conn.execute(
            "SELECT * FROM trades ORDER BY date DESC, id DESC"
        ).fetchall()
        conn.close()

        sells     = [t for t in trades if t["action"] == "SELL"]
        wins      = [t for t in sells  if t["pnl"] > 0]
        total_pnl = sum(t["pnl"] for t in sells)

        return jsonify({
            "trades":         [dict(t) for t in trades],
            "total_trades":   len(trades),
            "closed_trades":  len(sells),
            "winning_trades": len(wins),
            "win_rate":       round(len(wins) / max(1, len(sells)) * 100, 1),
            "total_pnl":      round(total_pnl, 2),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/trade-log")
def api_trade_log():
    """Return full auto-trade history: daily run summaries + individual trades."""
    with get_db() as con:
        runs = con.execute(
            "SELECT date, portfolio_value, cash, return_pct, trades_count, notes, created_at "
            "FROM daily_runs ORDER BY date DESC"
        ).fetchall()
        trades = con.execute(
            "SELECT date, action, ticker, name, sector, price, shares, pnl, signal, strategy, created_at "
            "FROM trades ORDER BY id DESC"
        ).fetchall()

    return jsonify({
        "daily_runs": [
            {
                "date":            r["date"],
                "portfolio_value": round(r["portfolio_value"], 2),
                "cash":            round(r["cash"], 2),
                "return_pct":      round(r["return_pct"], 4),
                "trades_count":    r["trades_count"],
                "notes":           r["notes"] or "",
                "created_at":      r["created_at"],
            }
            for r in runs
        ],
        "trades": [
            {
                "date":     t["date"],
                "action":   t["action"],
                "ticker":   t["ticker"],
                "name":     t["name"],
                "sector":   t["sector"],
                "price":    round(t["price"], 2),
                "shares":   t["shares"],
                "pnl":      round(t["pnl"], 2),
                "signal":   t["signal"],
                "strategy": t["strategy"],
                "time":     t["created_at"],
            }
            for t in trades
        ],
    })


@app.route("/api/run-now", methods=["POST"])
def api_run_now():
    """Manually trigger one auto-trade cycle (no need to wait for the scheduler)."""
    try:
        run_auto_trade()
        return jsonify({"ok": True, "message": "Auto-trade completed"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    init_db()
    start_scheduler()
    print("=" * 60)
    print("  AI Sector Paper Trading Simulator")
    print("  http://localhost:8080")
    print("=" * 60)
    # use_reloader=False 防止 APScheduler 被调度两次
    # host="0.0.0.0" 允许局域网/外网通过 IP 直接访问
    app.run(debug=False, host="0.0.0.0", port=8080, use_reloader=False)
